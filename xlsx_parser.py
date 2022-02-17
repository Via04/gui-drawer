from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class XLSXParser:
    MAX_LINE = 1000
    filepath = ''
    _ws = None
    _wb = None
    min_line = 1
    is_header = True

    def get_header(self, col: int) -> str:
        if self.is_header and self._ws is not None:
            val = self._ws.cell(1, col).value
            return val

    def get_column_values(self, col: int) -> tuple:
        res = []
        if self._ws is not None:
            col_generator = self._ws.iter_cols(min_col=col, max_col=col,
                                               min_row=self.min_line,
                                               max_row=self.MAX_LINE,
                                               values_only=True)  # generator
            res = [i for i in next(col_generator) if i is not None]
            return tuple(res)

    def __init__(self, filepath: str, is_header=True):
        self.filepath = filepath
        self.is_header = is_header
        if is_header:
            self.min_line = 2
        try:
            self._wb = load_workbook(filepath)  # loads .xlsx file
            self._wb.active = 0  # sets first sheet as active
            self._ws = self._wb.active  # a reference to a worksheet
        except (InvalidFileException, FileNotFoundError):
            print("Error! Bad path!")


# test = XLSXParser('ruave_out.xlsx')
# out = test.get_column_values(2)
# print(out)
# print(len(out))

