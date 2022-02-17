import PySimpleGUI as gui
import os
from os.path import curdir
# from xlsx_parser import XLSXParser
# from graph_drawer import PlotHelper
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from re import findall
import matplotlib.pyplot as plt
import numpy as np  # do not change np as it used as pattern addon
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from numba import njit
import time

def timeit(method):
    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()
        if 'log_time' in kw:
            name = kw.get('log_name', method.__name__.upper())
            kw['log_time'][name] = int((te - ts) * 1000)
        else:
            print(f'{method.__name__}  {(te - ts) * 1000}')
        return result
    return timed

PLUGIN_EXTENSION = 'py'
usr_model_layout = []
usr_model_functions = dict()
usr_common_layout = []
usr_common_functions = dict()

def analyze_plugs():
    is_model = False
    for files in os.listdir(os.getcwd()):
        if files.endswith(PLUGIN_EXTENSION):
            with open(files,'rt') as f:
                func_code = ''
                is_func = False
                user_action = ''
                for num, line in enumerate(f):
                    if num == 0 and line != '#!PLUGIN_SPECIFICATION':
                        break
                    if num == 1 and len(findall('\[MODEL_LAYOUT\]', line)) >= 1:
                        is_model = True
                    else:
                        is_model = False
                    if num == 2:
                        user_action = findall('\'.*\'', line)
                        user_action = user_action[1:-1] # убрать кавычки
                        user_layout_str = 'user_layout = ' + line
                        user_layout = None
                        _local = locals()
                        exec(user_layout_str, globals(), _local)
                        user_layout = _local['user_layout']
                        if is_model:
                            usr_model_layout.append(user_layout)
                        else:
                            usr_common_layout.append(user_layout)
                    if len(findall('Controller', line)) >= 1:
                        is_func = True
                        continue
                    if is_func:
                        func_code += line + '\n'
                func_code = '  '.join(('\n' + func_code.lstrip()).splitlines(True))
                func_code = f'def {files.split(".")[0]}(evt, val):\n' + func_code
                func_code = func_code + f'\nfunc_code_obj = {files.split(".")[0]}'
                func_code_obj = None
                _local = locals()
                exec(func_code, globals(), _local)
                func_code_obj = _local['func_code_obj']
                if is_model:
                    usr_model_functions.update({user_action: func_code_obj})
                else:
                    usr_common_functions.update({user_action: func_code_obj})
        else:
            continue

def draw_figure(canvas, figure):
    figure_canvas_agg = FigureCanvasTkAgg(figure, canvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
    return figure_canvas_agg

class XLSXParser:
    MAX_LINE = 1000
    filepath = ''
    _ws = None
    _wb = None
    min_line = 1
    is_header = True

    def get_header(self, col: int) -> str:
        if self.is_header and self._ws is not None:
            val = self._ws.cell(1, col).value
            return val

    def get_column_values(self, col: int) -> tuple:
        res = []
        if self._ws is not None:
            col_generator = self._ws.iter_cols(min_col=col, max_col=col,
                                               min_row=self.min_line,
                                               max_row=self.MAX_LINE,
                                               values_only=True)  # generator
            res = [i for i in next(col_generator) if i is not None]
            return tuple(res)

    def __init__(self, filepath: str, is_header=True):
        self.filepath = filepath
        self.is_header = is_header
        if is_header:
            self.min_line = 2
        try:
            self._wb = load_workbook(filepath)  # loads .xlsx file
            self._wb.active = 0  # sets first sheet as active
            self._ws = self._wb.active  # a reference to a worksheet
        except (InvalidFileException, FileNotFoundError):
            print("Error! Bad path!")


class PlotHelper:
    param_marker = ''
    val_marker = ''
    usr_func_marker = ''
    min_x = 0
    max_x = 0
    complexity = 0
    x = None
    fig = None
    ax = None

    def __init__(self, min_x=0, max_x=0, complexity=100, param_marker='p_', val_marker='v_',
                 usr_func_marker='u_'):
        self.param_marker = param_marker
        self.val_marker = val_marker
        self.usr_func_marker = usr_func_marker
        self.min_x = min_x
        self.max_x = max_x
        self.complexity = complexity
        self.x = np.linspace(min_x, max_x, complexity)
        self.fig = plt.figure(figsize=(6, 5))
        self.ax = self.fig.add_subplot(1, 1, 1)
    #@timeit
    def ans_expr(self, expr: str) -> list:
        code = self.gen_pycode(expr)
        x = self.x
        _local = locals()
        exec(code, globals(), _local)
        y = _local['y']
        return y

    def gen_pycode(self, expr: str):
        command = self.parse_expression(expr)
        y = None
        x = self.x
        # command = 'y = ' + command  # if you change y name variable, also change this line
        py_code = compile(command, '<string>', 'exec')
        return py_code

    def save_to_file(self, expr: str, filename):
        # expr = self.parse_expression(expr)
        with open(filename, 'w') as f:
            f.write(expr)

    def read_from_file(self, filename):
        with open(filename, 'r') as f:
            expr = ''.join(f.readlines())
            return expr

    def parse_expression(self, expr: str) -> str:
        final_expr = ''
        python_mode = False
        if expr is not None:
            for e in expr.split('\n'):
                if e != '':
                    if e[0] == '!':
                        python_mode = not python_mode
                        final_expr += '\n'
                        continue
                    if not python_mode:
                        # op_pattern = '\\s[A-XZa-xz]+\\s(?<!\\spow\\s)'
                        #  '\\s[^' + self.param_marker[:1] + self.val_marker[:1] + self.usr_func_marker[:1] + 'y' + \
                        #  '\\W\\d][^' + self.param_marker + self.val_marker + self.usr_func_marker + ']?\\w*\\s'
                        # op_pattern = '\s[^pv\W][^__]?\w*\s'
                        val_pattern = '\\s' + self.val_marker + '\\w*\\s'
                        param_pattern = '\\s' + self.param_marker + '\\w*\\s'
                        # ops = findall(op_pattern, expr)
                        vals = findall(val_pattern, e)
                        params = findall(param_pattern, e)
                        # for op in ops:
                        #     new_op = op[:1] + 'np.' + op[1:]
                        #     expr = expr.replace(op, new_op)
                        for val in vals:
                            new_val = ' ' + val[3:-1] + ' '
                            e = e.replace(val, new_val)
                        for param in params:
                            new_param = ' ' + param[3:-1] + ' '
                            e = e.replace(param, new_param)
                        expr_without_space = e.replace(' ', '')
                    else:
                        expr_without_space = e
                    final_expr += expr_without_space + '\n'
            # final_expr += 'return y'
            # final_expr = '  '.join(('\n' + final_expr.lstrip()).splitlines(True))
            # if is_parallel:
            #     final_expr = '@njit(parallel=True)\ndef myfunc(x):' + final_expr + '\ny = myfunc(x)'
            # else:
            #     final_expr = 'def myfunc(x):' + final_expr + '\ny = myfunc(x)'

            return final_expr

    def plot_expr(self, expr: str):
        y = self.ans_expr(expr)
        x = self.x
        plt.plot(x, y, 'ro-')
        plt.show()

    def get_text(self, expr: str) -> str:
        out = ''
        exprs = expr.split(';')
        for e in exprs:
            y = self.ans_expr(e)
            x = self.x
            for j in range(len(x)):
                print(f'{y[j]}\t{x[j]}\n')
                out += f'{y[j]}\t{x[j]}\n'
        return out

    def get_figure(self, expr: str):
        exprs = expr.split(';')
        for e in exprs:
            y = self.ans_expr(e)
            x = self.x
            ca = plt.gca()
            ca.set_ylim([min(y), max(y)])
            self.ax.grid(True)
            self.ax.plot(x, y, (self.max_x - self.min_x) / self.complexity)
        return self.fig

    @classmethod
    def plot_array(cls, x_list, y_list):
        plt.plot(x_list, y_list, 'ro-')
        plt.show()


layout = [[gui.In(key='input-file'), gui.FileBrowse(button_text='Выберите xlsx файл',
                                                    file_types=(('Excel files', '*.xlsx'),), target='input-file')],
          [gui.Button('Создание модели')],
          [gui.Button('Нарисовать график')]]
window = gui.Window('Тестирование результатов', layout)
file = window['input-file']
while True:
    event, values = window.read()
    if event == 'Нарисовать график':
        xlsx = XLSXParser(values['input-file'])
        x_list = xlsx.get_column_values(1)
        y_list = xlsx.get_column_values(2)
        PlotHelper.plot_array(x_list, y_list)
    if event == 'Создание модели':
        model_layout = [[gui.Multiline(key='-MODEL-', size=(100, 20), font='Monospaced 16'),
                         gui.Button('Create', visible=True)],
                        [gui.Button('Text table', visible=True)],
                        [gui.Text('Мин'), gui.In(key='-MIN-X-', size=(15, 10), default_text='0'),
                         gui.Text('Макс'), gui.In(key='-MAX-X-', size=(15, 10), default_text='100'),
                         gui.Text('Сложность'), gui.In(key='-COMPLEXITY-', size=(15, 10), default_text='100')],
                        [gui.InputText(key='-SAVE-MOD-', enable_events=True, visible=False),
                         gui.FileSaveAs('Save', target='-SAVE-MOD-', file_types=(('PYDATA', '.pydat'),),
                                        initial_folder=curdir),
                         gui.InputText(key='-LOAD-MOD-', enable_events=True, visible=False),
                         gui.FileBrowse('Load', target='-LOAD-MOD-', file_types=(('PYDATA', '.pydat'),),
                                        initial_folder=curdir)]]
        mod_window = gui.Window('Создайте модель', model_layout)
        while True:
            mod_evt, mod_val = mod_window.read()
            if mod_evt == 'Create':
                min_x = int(mod_val['-MIN-X-'])
                max_x = int(mod_val['-MAX-X-'])
                complexity = int(mod_val['-COMPLEXITY-'])
                plot = PlotHelper(min_x, max_x, complexity)
                model = mod_val['-MODEL-']
                plot_layout = [[gui.Text('Model')], [gui.Canvas(key='-CANVAS-')]]
                fig = plot.get_figure(model)
                canv_window = gui.Window('Graph from model', plot_layout, finalize=True,
                                         element_justification='center', font='Monospace 18')
                fig_agg = draw_figure(canv_window['-CANVAS-'].TKCanvas, fig)
                fig_agg.draw()
            if mod_evt == 'Text table':
                min_x = int(mod_val['-MIN-X-'])
                max_x = int(mod_val['-MAX-X-'])
                complexity = int(mod_val['-COMPLEXITY-'])
                plot = PlotHelper(min_x, max_x, complexity)
                model = mod_val['-MODEL-']
                out = plot.get_text(model)
                with open('out.txt', 'wt') as f:
                    f.write(out)
            if mod_evt == '-SAVE-MOD-':
                plot = PlotHelper()
                model = mod_val['-MODEL-']
                filename = mod_val['-SAVE-MOD-']
                plot.save_to_file(model, filename)
            if mod_evt == '-LOAD-MOD-':
                plot = PlotHelper()
                filename = mod_val['-LOAD-MOD-']
                if filename != '':
                    model = plot.read_from_file(filename)
                    mod_window['-MODEL-'].update(model)
            for key, func in usr_model_functions:
                if mod_evt == key:
                    func(mod_evt, mod_val)
            if mod_evt == gui.WINDOW_CLOSED:
                break
        break
    if event == gui.WINDOW_CLOSED:
        break

window.close()
